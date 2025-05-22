import frappe
from frappe import _
from frappe.utils import get_url, get_datetime
from frappe.utils.pdf import get_pdf
from frappe.utils.jinja import render_template
from frappe.utils.file_manager import save_file
from frappe import format_value
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


@frappe.whitelist()
def get_collection_pdf(site: str, date: str):
    """Return a File URL for the Collection-Details PDF.
       Skip any descendant location that has zero serialized assets.
       Abort entirely if *none* of the descendant locations contain assets.
    """
    site_doc = frappe.get_doc("Location", site)

    # 1️⃣ pull all descendants
    descendants = frappe.get_all(
        "Location",
        filters={"lft": (">=", site_doc.lft), "rgt": ("<=", site_doc.rgt)},
        order_by="lft asc",
        pluck="name",
    )

    # 2️⃣ collect assets, but keep only non-empty locations
    data = []
    for loc in descendants:
        assets = frappe.get_all(
            "Asset",
            filters={"location": loc},
            fields=["name as asset_name"],
            order_by="asset_name asc",
        )
        if assets:
            data.append({"location": loc, "assets": assets})

    if not data:
        frappe.throw(_("No serialized assets found under the selected site."))

    # 3️⃣ Format date for display
    formatted_date = format_value(get_datetime(date), {"fieldtype": "Date"})

    # 4️⃣ Render, convert, save
    context = {
        "data": data,
        "parent_location": site_doc.name,
        "date": formatted_date,
    }
    html = render_template("templates/collection_details_template.html", context)
    # pdf_bytes = get_pdf(html)
    pdf_bytes = get_pdf(
    html,
    options={
        "margin-bottom": "20mm",
        "footer-center": f"Page [page] of [topage] – {site_doc.name}",
        "footer-font-size": "9",
        "footer-spacing": "5"
    }
)


    file_doc = save_file(
        fname=f"Collection_Details_{site_doc.name}_{date}.pdf",
        content=pdf_bytes,
        dt=None,
        dn=None,
        folder="Home/Attachments",
        decode=False,
    )
    return file_doc.file_url


@frappe.whitelist()
def get_collection_excel(site: str, date: str):
    """Return a File URL for an Excel workbook grouped by descendant Locations."""
    site_doc = frappe.get_doc("Location", site)

    # ── collect descendant locations ──
    descendants = frappe.get_all(
        "Location",
        filters={"lft": (">=", site_doc.lft), "rgt": ("<=", site_doc.rgt)},
        order_by="lft asc",
        pluck="name",
    )
    print(descendants)

    # ── gather assets, skipping empties ──
    data = []
    for loc in descendants:
        assets = frappe.get_all(
            "Asset",
            filters={"location": loc},
            fields=["name as asset_name"],
            order_by="asset_name asc",
        )
        if assets:
            data.append({"location": loc, "assets": assets})

    if not data:
        frappe.throw(_("No serialized assets found under the selected site."))

    # ── build workbook ──
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    for loc in data:
        ws = wb.create_sheet(title=loc["location"][:31])  # sheet-name ≤31 chars
        ws.append(["#", "Asset", "Remarks"])
        for idx, asset in enumerate(loc["assets"], start=1):
            ws.append([idx, asset["asset_name"], ""])

        # basic column widths
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 30

    # ── save to bytes & File doctype ──
    bio = io.BytesIO()
    wb.save(bio)
    excel_bytes = bio.getvalue()

    file_doc = save_file(
        fname=f"Collection_Details_{site_doc.name}_{date}.xlsx",
        content=excel_bytes,
        dt=None,
        dn=None,
        folder="Home/Attachments",
        decode=False,
    )
    return file_doc.file_url





# import frappe
# from frappe.utils import get_url
# from frappe.utils.pdf import get_pdf
# from frappe.utils.jinja import render_template
# from frappe.utils.file_manager import save_file


# @frappe.whitelist()
# def get_collection_pdf(site: str, date: str):
#     """Return a File URL for the Collection-Details PDF.
#        Skip any descendant location that has zero serialized assets.
#        Abort entirely if *none* of the descendant locations contain assets.
#     """
#     site_doc = frappe.get_doc("Location", site)

#     # 1️⃣  pull all descendants
#     descendants = frappe.get_all(
#         "Location",
#         filters={"lft": (">", site_doc.lft), "rgt": ("<", site_doc.rgt)},
#         order_by="lft asc",
#         pluck="name",
#     )

#     # 2️⃣  collect assets, but keep only non-empty locations
#     data = []
#     for loc in descendants:
#         assets = frappe.get_all(
#             "Asset",
#             filters={"location": loc},
#             fields=["name as asset_name"],
#             order_by="asset_name asc",
#         )
#         if assets:                                 # ⬅️ keep only if something found
#             data.append({"location": loc, "assets": assets})

#     if not data:
#         frappe.throw(_("No serialized assets found under the selected site."))

#     # 3️⃣  render, convert, save (unchanged)
#     context = {"data": data, "parent_location": site_doc.name, "date": date}
#     html = frappe.render_template("templates/collection_details_template.html", context)
#     pdf_bytes = frappe.utils.pdf.get_pdf(html)
#     file_doc = save_file(
#         fname=f"Collection_Details_{site_doc.name}_{date}.pdf",
#         content=pdf_bytes,
#         dt=None,
#         dn=None,
#         folder="Home/Attachments",
#         decode=False,
#     )
#     return file_doc.file_url



# # your_app/collection_details.py  (add below the PDF function)
# import io
# # from frappe.utils.file_manager import save_file
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter

# @frappe.whitelist()
# def get_collection_excel(site: str, date: str):
#     """Return a File URL for an Excel workbook grouped by descendant Locations."""
#     site_doc = frappe.get_doc("Location", site)

#     # ── collect descendant locations ──
#     descendants = frappe.get_all(
#         "Location",
#         filters={"lft": (">", site_doc.lft), "rgt": ("<", site_doc.rgt)},
#         order_by="lft asc",
#         pluck="name",
#     )

#     # ── gather assets, skipping empties ──
#     data = []
#     for loc in descendants:
#         assets = frappe.get_all(
#             "Asset",
#             filters={"location": loc},
#             fields=["name as asset_name"],
#             order_by="asset_name asc",
#         )
#         if assets:
#             data.append({"location": loc, "assets": assets})

#     if not data:
#         frappe.throw(_("No serialized assets found under the selected site."))

#     # ── build workbook ──
#     wb = Workbook()
#     wb.remove(wb.active)          # drop the default sheet

#     for loc in data:
#         ws = wb.create_sheet(title=loc["location"][:31])  # sheet-name ≤31 chars
#         ws.append(["#", "Asset", "Remarks"])
#         for idx, asset in enumerate(loc["assets"], start=1):
#             ws.append([idx, asset["asset_name"], ""])

#         # basic column widths
#         for col in range(1, 4):
#             ws.column_dimensions[get_column_letter(col)].width = 30

#     # ── save to bytes & File doctype ──
#     bio = io.BytesIO()
#     wb.save(bio)
#     excel_bytes = bio.getvalue()



#     file_doc = save_file(
#         fname=f"Collection_Details_{site_doc.name}_{date}.xlsx",
#         content=excel_bytes,
#         dt=None,         # <-- no parent doctype
#         dn=None,         # <-- no parent docname
#         folder="Home/Attachments",
#         decode=False,
#     )
#     return file_doc.file_url

